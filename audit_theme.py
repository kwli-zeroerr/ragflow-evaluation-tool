import pandas as pd
from openpyxl import load_workbook

# 读取原始 Excel 文件
file_path = 'ZeroErrGPT_test.xlsx'  # 请将这里的路径替换为你自己的文件路径
wb = load_workbook(file_path)
ws = wb.active  # 获取活动的工作表

df = pd.read_excel(file_path)

# 定义theme列值的替换关系
theme_mapping = {
    'manual': 'eRob机器人关节模组用户手册_V3.39',
    'modbus': 'eRob机器人关节Modbus-RTU通信应用手册_V1.6',
    'ethercat': 'eRob CANopen and EtherCAT用户手册v1.9',
    'ecoder': 'eCoder编码器用户手册V2.4'
}

# 使用替换关系修改theme列的值
df['theme'] = df['theme'].map(theme_mapping).fillna(df['theme'])  # 如果没有匹配的值，保留原值

# 将修改后的数据写回工作表，保留原来的格式
for idx, row in df.iterrows():
    # 假设数据从第二行开始，第一行是标题
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=idx+2, column=col_idx, value=value)

# 保存为新的 Excel 文件
new_file_path = 'modified_file.xlsx'  # 修改后的文件路径
wb.save(new_file_path)

print(f"数据已经保存为 {new_file_path}")
